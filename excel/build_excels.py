#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Génère les 4 classeurs Excel Sia × Cosumar (branding navy/teal, gridlines off).
  - Scoring_Maturite_IA.xlsx      (8 dim. x 5 niveaux, radar auto)
  - Scoring_Maturite_Data.xlsx    (9 dim. x 6 niveaux, radar auto)
  - Grille_Audit_SI_Benchmark.xlsx(13 systèmes, grille SI, benchmark 14 opérateurs)
  - Matrice_Priorisation.xlsx     (scoring pondéré, ranking auto, bar chart)
Usage: python3 build_excels.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import RadarChart, BarChart, Reference
from openpyxl.utils import get_column_letter

OUT = os.path.dirname(os.path.abspath(__file__))

# ---- Sia palette ----
NAVY   = "1F4E79"; NAVY_D = "0D2B45"; NAVY2 = "2E75B6"
TEAL   = "00A896"; GOLD   = "C9972B"; CREAM = "F7F9FC"
WHITE  = "FFFFFF"; GREY   = "6B7A99"; LINE  = "D8E0EC"

F_TITLE  = Font(name="Calibri", size=18, bold=True, color=WHITE)
F_SUB    = Font(name="Calibri", size=10, color="DCE6F2")
F_H      = Font(name="Calibri", size=10, bold=True, color=WHITE)
F_DIM    = Font(name="Calibri", size=10, bold=True, color=NAVY_D)
F_TXT    = Font(name="Calibri", size=10, color="1A2744")
F_MUT    = Font(name="Calibri", size=9, color=GREY)
F_BIG    = Font(name="Calibri", size=22, bold=True, color=NAVY)

FILL_TITLE = PatternFill("solid", fgColor=NAVY_D)
FILL_BAND  = PatternFill("solid", fgColor=NAVY)
FILL_HEAD  = PatternFill("solid", fgColor=NAVY2)
FILL_TEAL  = PatternFill("solid", fgColor=TEAL)
FILL_CREAM = PatternFill("solid", fgColor=CREAM)
FILL_GOLD  = PatternFill("solid", fgColor=GOLD)
FILL_WHITE = PatternFill("solid", fgColor=WHITE)

CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEF = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin = Side(style="thin", color=LINE)
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)


def base_sheet(ws, title, subtitle, ncols=8):
    ws.sheet_view.showGridLines = False
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(1, 1, "  SIA  ·  COSUMAR")
    c.font = Font(name="Calibri", size=11, bold=True, color=TEAL)
    c.fill = FILL_TITLE; c.alignment = LEF
    for col in range(1, ncols + 1):
        ws.cell(1, col).fill = FILL_TITLE; ws.cell(2, col).fill = FILL_TITLE
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
    t = ws.cell(3, 1, title); t.font = F_TITLE; t.fill = FILL_BAND; t.alignment = LEF
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=ncols)
    s = ws.cell(4, 1, subtitle); s.font = F_SUB; s.fill = FILL_BAND; s.alignment = LEF
    ws.row_dimensions[1].height = 18; ws.row_dimensions[3].height = 30; ws.row_dimensions[4].height = 18
    for col in range(1, ncols + 1):
        ws.cell(3, col).fill = FILL_BAND; ws.cell(4, col).fill = FILL_BAND


def header_row(ws, row, headers, widths=None):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row, i, h); c.font = F_H; c.fill = FILL_HEAD; c.alignment = CEN; c.border = BORD
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 28


# ============================================================
# 1) SCORING MATURITE IA — 8 dimensions x 5 niveaux
# ============================================================
def build_maturite_ia():
    wb = Workbook(); ws = wb.active; ws.title = "Scoring Maturité IA"
    base_sheet(ws, "Scoring de Maturité IA",
               "8 dimensions × 5 niveaux · Exploratoire ▸ Émergent ▸ Structuré ▸ Avancé ▸ Transformé", 6)
    dims = [
        ("Stratégie & Vision IA", "Ambition IA portée par la direction, alignée au business"),
        ("Données & Infrastructure", "Qualité, accessibilité, plateforme data"),
        ("Talents & Compétences", "Profils data/IA, formation, attractivité"),
        ("Gouvernance & Éthique", "Cadre de décision, conformité CNDP, IA responsable"),
        ("Cas d'usage & Valeur", "Portefeuille de UC, ROI démontré"),
        ("Technologie & MLOps", "Outils, industrialisation, cycle de vie modèles"),
        ("Culture & Adoption", "Acculturation, conduite du changement"),
        ("Écosystème & Partenariats", "Partenaires tech, recherche, startups"),
    ]
    levels = ["Exploratoire", "Émergent", "Structuré", "Avancé", "Transformé"]
    r = 6
    head = ["Dimension", "Description", "Niveau actuel (1-5)", "Cible (1-5)", "Écart", "Commentaire"]
    header_row(ws, r, head, [26, 40, 16, 12, 9, 30]); r += 1
    sample_now = [3, 2, 2, 3, 2, 2, 3, 2]
    sample_tgt = [4, 4, 4, 4, 5, 4, 4, 3]
    first = r
    for i, (d, desc) in enumerate(dims):
        ws.cell(r, 1, d).font = F_DIM
        ws.cell(r, 2, desc).font = F_TXT
        ws.cell(r, 3, sample_now[i]).font = F_TXT
        ws.cell(r, 4, sample_tgt[i]).font = F_TXT
        ws.cell(r, 5, f"=D{r}-C{r}").font = Font(bold=True, color=GOLD)
        ws.cell(r, 6, "").font = F_TXT
        fill = FILL_WHITE if i % 2 == 0 else FILL_CREAM
        for col in range(1, 7):
            ws.cell(r, col).fill = fill; ws.cell(r, col).border = BORD
            ws.cell(r, col).alignment = CEN if col in (3, 4, 5) else LEF
        ws.row_dimensions[r].height = 30
        r += 1
    last = r - 1
    # score global
    ws.cell(r, 1, "Score global IA (moyenne /5)").font = F_DIM
    ws.cell(r, 3, f"=ROUND(AVERAGE(C{first}:C{last}),1)").font = F_BIG
    ws.cell(r, 4, f"=ROUND(AVERAGE(D{first}:D{last}),1)").font = Font(size=14, bold=True, color=TEAL)
    for col in range(1, 7):
        ws.cell(r, col).fill = FILL_CREAM; ws.cell(r, col).border = BORD
    ws.cell(r, 3).alignment = CEN; ws.cell(r, 4).alignment = CEN
    ws.row_dimensions[r].height = 34

    # Échelle de lecture
    r += 2
    ws.cell(r, 1, "Échelle des niveaux").font = F_DIM; r += 1
    for i, lv in enumerate(levels, 1):
        ws.cell(r, 1, i).font = F_H; ws.cell(r, 1).fill = FILL_HEAD; ws.cell(r, 1).alignment = CEN; ws.cell(r, 1).border = BORD
        ws.cell(r, 2, lv).font = F_DIM; ws.cell(r, 2).border = BORD
        r += 1

    # Radar
    chart = RadarChart(); chart.type = "filled"; chart.style = 26
    chart.title = "Profil de maturité IA — Actuel vs Cible"
    data = Reference(ws, min_col=3, max_col=4, min_row=first - 1, max_row=last)
    cats = Reference(ws, min_col=1, min_row=first, max_row=last)
    chart.add_data(data, titles_from_data=True); chart.set_categories(cats)
    chart.height = 11; chart.width = 16
    ws.add_chart(chart, "H6")
    wb.save(os.path.join(OUT, "Scoring_Maturite_IA.xlsx"))
    print("✓ Scoring_Maturite_IA.xlsx")


# ============================================================
# 2) SCORING MATURITE DATA — 9 dimensions x 6 niveaux
# ============================================================
def build_maturite_data():
    wb = Workbook(); ws = wb.active; ws.title = "Scoring Maturité Data"
    base_sheet(ws, "Scoring de Maturité Data",
               "9 dimensions × 6 niveaux · 35 questions · Inexistant ▸ Initial ▸ Reproductible ▸ Défini ▸ Géré ▸ Optimisé", 6)
    dims = [
        ("Stratégie Data", 4), ("Gouvernance Data", 4), ("Qualité des données", 4),
        ("Architecture & Stockage", 4), ("Intégration & Pipelines", 4), ("Analytics & BI", 4),
        ("Sécurité & Conformité", 4), ("Culture Data", 3), ("Data Ops & Outils", 4),
    ]  # total questions = 35
    levels = ["Inexistant", "Initial", "Reproductible", "Défini", "Géré", "Optimisé"]
    r = 6
    header_row(ws, r, ["Dimension", "Nb questions", "Niveau actuel (0-5)", "Cible (0-5)", "Écart", "Commentaire"],
               [28, 14, 18, 12, 9, 30]); r += 1
    sample_now = [3, 2, 2, 3, 2, 3, 3, 2, 2]
    sample_tgt = [5, 4, 4, 4, 4, 4, 5, 4, 4]
    first = r
    for i, (d, nq) in enumerate(dims):
        ws.cell(r, 1, d).font = F_DIM
        ws.cell(r, 2, nq).font = F_TXT
        ws.cell(r, 3, sample_now[i]).font = F_TXT
        ws.cell(r, 4, sample_tgt[i]).font = F_TXT
        ws.cell(r, 5, f"=D{r}-C{r}").font = Font(bold=True, color=GOLD)
        fill = FILL_WHITE if i % 2 == 0 else FILL_CREAM
        for col in range(1, 7):
            ws.cell(r, col).fill = fill; ws.cell(r, col).border = BORD
            ws.cell(r, col).alignment = CEN if col in (2, 3, 4, 5) else LEF
        ws.row_dimensions[r].height = 26
        r += 1
    last = r - 1
    ws.cell(r, 1, "Score global Data (/5)").font = F_DIM
    ws.cell(r, 2, f"=SUM(B{first}:B{last})").font = F_MUT
    ws.cell(r, 3, f"=ROUND(AVERAGE(C{first}:C{last}),1)").font = F_BIG
    ws.cell(r, 4, f"=ROUND(AVERAGE(D{first}:D{last}),1)").font = Font(size=14, bold=True, color=TEAL)
    for col in range(1, 7):
        ws.cell(r, col).fill = FILL_CREAM; ws.cell(r, col).border = BORD
    ws.cell(r, 3).alignment = CEN; ws.cell(r, 4).alignment = CEN
    ws.row_dimensions[r].height = 32

    r += 2
    ws.cell(r, 1, "Échelle des niveaux (0-5)").font = F_DIM; r += 1
    for i, lv in enumerate(levels):
        ws.cell(r, 1, i).font = F_H; ws.cell(r, 1).fill = FILL_HEAD; ws.cell(r, 1).alignment = CEN; ws.cell(r, 1).border = BORD
        ws.cell(r, 2, lv).font = F_DIM; ws.cell(r, 2).border = BORD
        r += 1

    chart = RadarChart(); chart.type = "filled"; chart.style = 18
    chart.title = "Profil de maturité Data — Actuel vs Cible"
    data = Reference(ws, min_col=3, max_col=4, min_row=first - 1, max_row=last)
    cats = Reference(ws, min_col=1, min_row=first, max_row=last)
    chart.add_data(data, titles_from_data=True); chart.set_categories(cats)
    chart.height = 11; chart.width = 16
    ws.add_chart(chart, "H6")
    wb.save(os.path.join(OUT, "Scoring_Maturite_Data.xlsx"))
    print("✓ Scoring_Maturite_Data.xlsx")


# ============================================================
# 3) GRILLE AUDIT SI & BENCHMARK
# ============================================================
def build_audit_si():
    wb = Workbook()
    # --- Sheet 1: Inventaire SI ---
    ws = wb.active; ws.title = "Inventaire SI"
    base_sheet(ws, "Audit du Système d'Information", "Inventaire des 13 systèmes & évaluation (10 critères)", 7)
    systems = [
        ("SAP S/4 HANA", "ERP cœur (MM, PP, PM, QM, CO, FI, SD, LE)", "Finance / Industrie / SC"),
        ("SAP Ariba", "Achats & sourcing", "Achats"),
        ("Salesforce / C4C", "CRM & relation client", "Commerciale"),
        ("Qubes MES", "Manufacturing Execution System", "Industrielle"),
        ("SCERI", "Gestion agricole amont", "Amont Agricole"),
        ("HR Access", "SIRH", "RH"),
        ("SAP IBP", "Integrated Business Planning", "Supply Chain"),
        ("SAP BPC", "Planification & consolidation", "Finance"),
        ("Business Objects", "Reporting BI", "Transversal"),
        ("Power BI", "Dataviz & dashboards", "Transversal"),
        ("Attaissir", "Solution métier locale", "Transversal"),
        ("Data Lake / DWH", "Stockage analytique", "SI / Data"),
        ("Portail collaboratif", "Intranet & GED", "SI"),
    ]
    r = 6
    header_row(ws, r, ["Système", "Description", "Direction", "Criticité (1-5)",
                       "Qualité data (1-5)", "Ouverture API (1-5)", "Potentiel IA (1-5)"],
               [20, 36, 20, 13, 13, 13, 13]); r += 1
    crit = [5,3,4,5,4,3,4,3,3,4,2,4,2]
    qual = [4,3,3,4,3,3,4,4,3,4,2,3,3]
    api  = [4,3,4,3,2,2,4,3,2,4,2,4,3]
    pot  = [5,3,4,5,4,3,5,3,2,4,2,5,2]
    first = r
    for i, (s, d, dr) in enumerate(systems):
        ws.cell(r, 1, s).font = F_DIM
        ws.cell(r, 2, d).font = F_TXT
        ws.cell(r, 3, dr).font = F_MUT
        for col, arr in zip((4,5,6,7), (crit, qual, api, pot)):
            ws.cell(r, col, arr[i]).font = F_TXT
        fill = FILL_WHITE if i % 2 == 0 else FILL_CREAM
        for col in range(1, 8):
            ws.cell(r, col).fill = fill; ws.cell(r, col).border = BORD
            ws.cell(r, col).alignment = CEN if col >= 4 else LEF
        ws.row_dimensions[r].height = 26
        r += 1
    last = r - 1
    ws.cell(r, 1, "Moyenne").font = F_DIM
    for col in (4,5,6,7):
        L = get_column_letter(col)
        ws.cell(r, col, f"=ROUND(AVERAGE({L}{first}:{L}{last}),1)").font = Font(bold=True, color=TEAL)
        ws.cell(r, col).alignment = CEN
    for col in range(1, 8):
        ws.cell(r, col).fill = FILL_CREAM; ws.cell(r, col).border = BORD

    # bar chart potentiel IA
    chart = BarChart(); chart.type = "bar"; chart.style = 10; chart.title = "Potentiel IA par système"
    data = Reference(ws, min_col=7, min_row=first - 1, max_row=last)
    cats = Reference(ws, min_col=1, min_row=first, max_row=last)
    chart.add_data(data, titles_from_data=True); chart.set_categories(cats)
    chart.height = 10; chart.width = 18; chart.legend = None
    ws.add_chart(chart, "A" + str(r + 2))

    # --- Sheet 2: Benchmark ---
    ws2 = wb.create_sheet("Benchmark Opérateurs")
    base_sheet(ws2, "Benchmark sectoriel — opérateurs sucriers", "14 opérateurs · maturité IA & Data (score /5)", 6)
    ops = [
        ("Tereos", "France", 3.5, 3.2, "Optimisation agronomique, IoT industriel"),
        ("Südzucker", "Allemagne", 3.8, 3.6, "Maintenance prédictive, jumeau numérique"),
        ("Al Khaleej", "Arabie Saoudite", 2.8, 2.6, "Automatisation raffinage"),
        ("Illovo Sugar", "Afrique du Sud", 2.9, 2.7, "Prévision récolte, télédétection"),
        ("Cristal Union", "France", 3.3, 3.0, "Demand forecasting, énergie"),
        ("Nordzucker", "Allemagne", 3.6, 3.4, "Analytics avancé, durabilité"),
        ("ASR Group", "USA", 3.4, 3.1, "Supply chain IA, traçabilité"),
        ("Wilmar Sugar", "Singapour", 3.0, 2.9, "Optimisation logistique"),
        ("Mitr Phol", "Thaïlande", 3.1, 2.8, "Agriculture de précision"),
        ("Raízen", "Brésil", 3.7, 3.5, "Bioénergie, optimisation process"),
        ("Dangote Sugar", "Nigeria", 2.5, 2.3, "Digitalisation amont"),
        ("Kenana Sugar", "Soudan", 2.4, 2.2, "Gestion irrigation"),
        ("Cosumar (cible)", "Maroc", 2.6, 2.5, "Roadmap IA en cours — Sia"),
        ("Moyenne secteur", "—", 3.1, 2.9, "Référence comparative"),
    ]
    r = 6
    header_row(ws2, r, ["Opérateur", "Pays", "Maturité IA (/5)", "Maturité Data (/5)", "Initiatives phares"],
               [22, 18, 15, 16, 44]); r += 1
    for i, (n, p, mia, md, ini) in enumerate(ops):
        cos = "Cosumar" in n
        ws2.cell(r, 1, n).font = Font(bold=True, color=(TEAL if cos else NAVY_D), size=10)
        ws2.cell(r, 2, p).font = F_MUT
        ws2.cell(r, 3, mia).font = F_TXT
        ws2.cell(r, 4, md).font = F_TXT
        ws2.cell(r, 5, ini).font = F_TXT
        fill = PatternFill("solid", fgColor="E6F7F4") if cos else (FILL_WHITE if i % 2 == 0 else FILL_CREAM)
        for col in range(1, 6):
            ws2.cell(r, col).fill = fill; ws2.cell(r, col).border = BORD
            ws2.cell(r, col).alignment = CEN if col in (2,3,4) else LEF
        ws2.row_dimensions[r].height = 24
        r += 1
    chart2 = BarChart(); chart2.type = "col"; chart2.style = 12; chart2.title = "Maturité IA vs Data — Benchmark"
    data = Reference(ws2, min_col=3, max_col=4, min_row=6, max_row=r - 1)
    cats = Reference(ws2, min_col=1, min_row=7, max_row=r - 1)
    chart2.add_data(data, titles_from_data=True); chart2.set_categories(cats)
    chart2.height = 9; chart2.width = 22
    ws2.add_chart(chart2, "A" + str(r + 2))

    wb.save(os.path.join(OUT, "Grille_Audit_SI_Benchmark.xlsx"))
    print("✓ Grille_Audit_SI_Benchmark.xlsx")


# ============================================================
# 4) MATRICE DE PRIORISATION
# ============================================================
def build_priorisation():
    wb = Workbook(); ws = wb.active; ws.title = "Matrice Priorisation"
    base_sheet(ws, "Matrice de priorisation des Use Cases IA",
               "Scoring multicritère pondéré · ranking automatique", 9)
    # weights
    weights = {"Valeur métier": 0.30, "Faisabilité": 0.25, "Complexité inv.": 0.15,
               "Impact": 0.20, "Timeline": 0.10}
    r = 6
    ws.cell(r, 1, "Pondérations →").font = F_DIM
    cstart = 4
    for i, (k, w) in enumerate(weights.items()):
        c = ws.cell(r, cstart + i, f"{int(w*100)}%"); c.font = Font(bold=True, color=GOLD); c.alignment = CEN
        c.fill = FILL_GOLD; c.font = Font(bold=True, color=WHITE); c.border = BORD
    ws.row_dimensions[r].height = 20
    r += 1
    head = ["Rang", "Use Case", "Direction"] + list(weights.keys()) + ["Score pondéré"]
    header_row(ws, r, head, [8, 30, 18, 13, 12, 14, 11, 12, 15]); r += 1
    # ucs: scores 1-5 per criterion (complexité: 5 = simple/faible inv.)
    ucs = [
        ("OCR factures fournisseurs", "Financière", 4, 5, 5, 4, 5),
        ("Demand forecasting sucre", "Supply Chain", 5, 4, 4, 4, 5),
        ("Détection anomalies énergie", "Industrielle", 4, 4, 4, 4, 5),
        ("Chatbot RH CorpGPT", "RH", 4, 4, 3, 4, 4),
        ("Maintenance prédictive turbines", "Industrielle", 5, 3, 3, 5, 3),
        ("Route optimization livraisons", "Supply Chain", 4, 3, 3, 4, 4),
        ("Analyse contrats NLP", "Juridique", 3, 3, 4, 3, 3),
        ("Prévision récolte satellite", "Amont Agricole", 4, 2, 2, 4, 2),
        ("Optimisation process sucrier", "Industrielle", 5, 2, 2, 5, 2),
        ("People analytics & turnover", "RH", 3, 3, 3, 3, 2),
    ]
    first = r
    wlist = list(weights.values())
    for i, (n, d, *sc) in enumerate(ucs):
        ws.cell(r, 2, n).font = F_DIM
        ws.cell(r, 3, d).font = F_MUT
        for j, v in enumerate(sc):
            ws.cell(r, 4 + j, v).font = F_TXT; ws.cell(r, 4 + j).alignment = CEN
        # weighted score formula
        ws.cell(r, 9, f"=ROUND(D{r}*{wlist[0]}+E{r}*{wlist[1]}+F{r}*{wlist[2]}+G{r}*{wlist[3]}+H{r}*{wlist[4]},2)").font = Font(bold=True, color=NAVY)
        ws.cell(r, 9).alignment = CEN
        # rank
        ws.cell(r, 1, f"=RANK(I{r},$I${first}:$I${first+len(ucs)-1})").font = Font(bold=True, color=TEAL)
        ws.cell(r, 1).alignment = CEN
        fill = FILL_WHITE if i % 2 == 0 else FILL_CREAM
        for col in range(1, 10):
            ws.cell(r, col).fill = fill; ws.cell(r, col).border = BORD
            if col in (2, 3) and col == 2:
                ws.cell(r, col).alignment = LEF
        ws.cell(r, 2).alignment = LEF; ws.cell(r, 3).alignment = LEF
        ws.row_dimensions[r].height = 24
        r += 1
    last = r - 1

    # bar chart of scores
    chart = BarChart(); chart.type = "bar"; chart.style = 11; chart.title = "Scores pondérés des use cases"
    data = Reference(ws, min_col=9, min_row=first - 1, max_row=last)
    cats = Reference(ws, min_col=2, min_row=first, max_row=last)
    chart.add_data(data, titles_from_data=True); chart.set_categories(cats)
    chart.height = 11; chart.width = 20; chart.legend = None
    ws.add_chart(chart, "A" + str(r + 2))

    # --- Methodologie sheet ---
    ws2 = wb.create_sheet("Méthodologie")
    base_sheet(ws2, "Méthodologie de priorisation", "Funnel 3 filtres & cadre Build/Buy/Borrow/Bot", 4)
    rows = [
        ("Filtre 1 — Présélection collaborative", "40-60 idées → 20-25 UC. Ateliers participatifs, vote, alignement stratégique."),
        ("Filtre 2 — Qualification technique / ROI", "Faisabilité data/tech, business case, choix Build/Buy/Borrow/Bot."),
        ("Filtre 3 — Priorisation COPIL multicritère", "Scoring pondéré ci-contre, arbitrage COPIL, séquencement en 3 vagues."),
        ("", ""),
        ("Critère", "Pondération & sens"),
        ("Valeur métier", "30% — gain business attendu"),
        ("Faisabilité", "25% — maturité data, tech, compétences"),
        ("Complexité d'investissement", "15% — note inversée (5 = faible coût)"),
        ("Impact", "20% — portée, transformation, stratégie"),
        ("Timeline", "10% — rapidité de mise en œuvre"),
        ("", ""),
        ("Build", "Développement interne — différenciant, données propriétaires"),
        ("Buy", "Solution éditeur — besoin standard, time-to-market"),
        ("Borrow", "Partenaire / API — expertise externe, capacité ponctuelle"),
        ("Bot", "Automatisation / RPA — tâches répétitives à règles"),
    ]
    r = 6
    for i, (a, b) in enumerate(rows):
        if a in ("Critère",) or a.startswith("Filtre") or a in ("Build","Buy","Borrow","Bot"):
            ws2.cell(r, 1, a).font = F_DIM
        else:
            ws2.cell(r, 1, a).font = F_TXT
        ws2.cell(r, 2, b).font = F_TXT
        ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        ws2.cell(r, 1).alignment = LEF; ws2.cell(r, 2).alignment = LEF
        if a:
            for col in range(1, 5):
                ws2.cell(r, col).border = BORD
        ws2.row_dimensions[r].height = 26
        r += 1
    ws2.column_dimensions["A"].width = 32
    for col in "BCD":
        ws2.column_dimensions[col].width = 24

    wb.save(os.path.join(OUT, "Matrice_Priorisation.xlsx"))
    print("✓ Matrice_Priorisation.xlsx")


if __name__ == "__main__":
    build_maturite_ia()
    build_maturite_data()
    build_audit_si()
    build_priorisation()
    print("\nTous les classeurs Excel ont été générés dans:", OUT)
